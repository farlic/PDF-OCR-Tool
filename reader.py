import re,pdf2image,PIL,pytesseract,time,os,math,sys,regex,openpyxl,csv,time
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
os.system('mode 50,40')

fields=[
    "id",
	"clientref",
	"clientref2",
	"dforename",
	"dmiddlename1",
    "dmiddlename2",
	"dmiddlename3",
	"dsurname",
	"ddob",
	"daliasname",
    "dmaidenname",
	"dplaceofbirth",
	"daddress1",
	"daddress2",
	"daddress3",
    "daddress4",
	"daddress5",
	"dpostcode",
	"ddate",
	"district",
    "country",
	"drireference",
	"volume",
	"page",
	"regno",
    "entry",
	"districtno",
	"dor",
	"matchsource",
	"matchtype",
    "matchaudit",
	"matchnotes",
	"documentsource",
	"documentaddress",
	"contacttitle",
    "contactforename",
	"contactmiddlename",
	"contactsurname",
	"contactaddress1",
	"contactaddress2",
    "contactaddress3",
	"contactaddress4",
	"contactaddress5",
	"contactpostcode",
	"contacttelephone",
    "contactrelationshiptosubject",
	"contactresultcode"]
class Person:
    def __init__(self):
        self.district = ''
        self.dod = ''
        self.name = ''
        self.midname1 = ''
        self.midname2 = ''
        self.midname3 = ''
        self.surname = ''
        self.alias = ''
        self.gender = ''
        self.maiden = ''
        self.dob = ''
        self.pob = ''
        self.usual_address = ''
        self.informant_name = ''
        self.informant_midname = ''
        self.informant_surname = ''
        self.qualification = ''
        self.informant_address = ''
        self.reg_date = ''
        self.dri = ''
error_list = []
pos_data = {
    "district":20,
    "dod":19,
    "name":4,
    "midname1":5,
    "midname2":6,
    "midname3":7,
    "surname":8,
    "alias":10,
    "maiden":11,
    "dob":9,
    "pob":12,
    "usual_address":13,
    "informant_name":36,
    "informant_midname":37,
    "informant_surname":38,
    "qualification":46,
    "informant_address":39,
    "reg_date":28,
    "dri":22}
days ={
    "first":"01",
    "second":"02",
    "third":"03",
    "fourth":"04",
    "fifth":"05",
    "sixth":"06",
    "seventh":"07",
    "eighth":"08",
    "ninth":"09",
    "tenth":"10",
    "eleventh":"11",
    "twelfth":"12",
    "thirteenth":"13",
    "fourteenth":"14",
    "fifteenth":"15",
    "sixteenth":"16",
    "seventeenth":"17",
    "eighteenth":"18",
    "ninteenth":"19",
    "twentieth":"20",
    "twenty-first":"21",
    "twenty-second":"22",
    "twenty-third":"23",
    "twenty-fourth":"24",
    "twenty-fifth":"25",
    "twenty-sixth":"26",
    "twenty-seventh":"27",
    "twenty-eighth":"28",
    "twenty-ninth":"29",
    "thirtieth":"30",
    "thirty-first":"31"}
months = {
    "january":"01",
    "february":"02",
    "march":"03",
    "april":"04",
    "may":"05",
    "june":"06",
    "july":"07",
    "august":"08",
    "september":"09",
    "october":"10",
    "november":"11",
    "december":"12"}

def convert_month(month_string:str):
    return months[month_string]

def convert_day(day_string:str):
    return days[day_string]

def convert_date(date_string:str):
    date = date_string.split(' ')
    return f'{convert_day(date[0])}/{convert_month(date[1])}/{date[-1]}'

def title():
    print(r'''
   _____  _____  ______    ____   _____ _____  
  |  __ \|  __ \|  ____|  / __ \ / ____|  __ \ 
  | |__) | |  | | |__    | |  | | |    | |__) |
  |  ___/| |  | |  __|   | |  | | |    |  _  / 
  | |    | |__| | |      | |__| | |____| | \ \ 
  |_|    |_____/|_|       \____/ \_____|_|  \_\
                                            
    ''')
    print(f'--------------------------------------------------')    
title()

print('setting working directory')
pytesseract.pytesseract.tesseract_cmd = os.getcwd()+'//Tesseract-OCR//tesseract'
print('pytesseract located\nloading PDF')
print(f'--------------------------------------------------')
for file__ in os.listdir():
    if file__.endswith('pdfocr.csv'):
        os.remove(file__)
        print('old output purged')

pdfs = [file_ for file_ in os.listdir() if file_.endswith(".pdf") or file_.endswith('.PDF')]
if len(pdfs) >1:
	pdf = None
	print('multiple PDFs detected: please select one')
	for key,files in enumerate(pdfs):
		print(f"[{key}]: {files.split('.')[0]}")
	while not pdf:
		try:	
			choice = input('please select ... ')
			pdf = pdfs[int(choice)]
		except ValueError:
			print('select using a number')
else:
	pdf = pdfs[0]
pages = pdf2image.convert_from_path(pdf,poppler_path=os.getcwd()+'//poppler-0.68.0//bin') #grab page objects
print('pdf loaded')

wb = Workbook()
ws = wb.active
ws.title = "data"

print('--------------------------------------------------')
print('compiling regular expressions')

gender_pattern = re.compile("( (fe)?male)",re.I)
relative_pattern = re.compile("((bro|mo|fa)?ther|sister|informant|causing the body to be cremated|uncle)",re.I)
district_pattern = regex.compile('(?i)(registration district|Administrative area){s<=2,d<=2,e<=3}')
time.sleep(1)

for index,page in enumerate(pages):
    p = Person()

    try:
        start_time = time.time()
        
        os.system('cls')
        title()
        print(f'processing page {index+1}')

        text = pytesseract.image_to_string(page)
        #grab text
        print('text found')

        bad = [",",".",")","(","'"]
        for x in bad:
            text = text.replace(x,"")

        text = text.split('\n')
        text = [line for line in text if line.strip() !='']
        #split into lines, strip empty lines

        for key,line in enumerate(text): 
            #print(line) #temp output to see lines

            if regex.search('(?i)(registration district){e<=2}',line) and len(p.district) == 0: #(?i) is the old regex ignore case
                p.district = text[key].strip()

            elif regex.search('(?i)(date and place of death){e<=3}',line) and len(p.dod) == 0:# {e<3} is fuzzy match
                p.dod = text[key+1].strip()

            elif regex.search('(?i)(name and surname){e<=2}',line) and len(p.name) == 0:
                p.name = text[key+1].strip()

            elif regex.search('(?i)(of woman who){e<=2}',line) and len(p.maiden) == 0:
                p.maiden = text[key].strip().split(' ')[-1]

            elif regex.search('(?i)(date and place of birth){e<=3}',line) and len(p.dob) == 0:
                p.dob = text[key+1].strip()
                p.pob = text[key+2].strip()

            #elif regex.search('(?i)(date and place of birth){e<=3}',line) and len(p.pob) == 0:
            #    p.pob = text[key+2].strip()

            elif regex.search('(?i)(name and surname of informant){e<=3}',line) and len(p.usual_address) == 0:
                p.usual_address = text[key-1].strip()

            #elif regex.search('(?i)(informant){e<=3}',line) and len(p.informant_name) == 0:
            #    p.informant_name = text[key+1].strip()

            elif regex.search('(?i)(certify that){e<=3}',line) and len(p.informant_address) == 0:
                p.informant_address = text[key-1].strip()
                p.informant_name = text[key-3].strip()

                if regex.search('(?i)(present at death){e<=3}',p.informant_name):
                    p.informant_name = text[key-4].strip()

            elif regex.search('(?i)(date of registration){e<=3}',line) and len(p.reg_date) == 0:
                p.reg_date = text[key+1].strip()

            elif regex.search('(?i)(system no){e<=3}',line) and len(p.dri) == 0:
                p.dri = text[key][regex.search(r'(\d){3,}',line).span()[0]:regex.search(r'(\d){3,}',line).span()[1]+1].strip()
            
        #cleaning scraped results

        p.gender = gender_pattern.search(p.name[-7:]).group(0).strip()

        p.name = p.name[:-7]+gender_pattern.sub("",p.name[-7:])
        names = []
        names = p.name.split(' ')
        p.name = names[0].strip()
        p.surname = names[-1].strip()
        if len(names[1:-1]) >3:
            ' '.join(names[3:-1])
            names[4:-1] = ''
        for names_key,remaining_names in enumerate(names[1:-1]):
            setattr(p,f"midname{names_key+1}",remaining_names.strip()) 

        p.reg_date = p.reg_date[:re.search(r'\d+',p.reg_date).span()[1]]
        p.reg_date = convert_month(p.reg_date.split(' ')[1].lower())+'-'+p.reg_date.split(' ')[-1]

        p.dob = convert_date(p.dob.lower())
        p.dod = convert_date(p.dod.lower())

        if len(p.pob.split(' ')[-1]) == 1:
            p.pob = ' '.join(p.pob.split(' ')[:-1])

        remove_districts = []
        district_iterator = district_pattern.finditer(p.district)
        for match in district_iterator:
            remove_districts.append(match.span())
        p.district = p.district[remove_districts[0][1]:remove_districts[1][0]].strip()

        if p.gender.lower() == 'male':
            p.maiden = ''

        qualification_span = []
        qualification_span = regex.search(r'([A-Z]){2,}\s{s<=2}',p.informant_name).span()
        p.qualification = p.informant_name[qualification_span[1]:]
        
        p.informant_name = p.informant_name[:qualification_span[1]-1]
        informant_names = []
        informant_names = p.informant_name.split(' ')
        p.informant_name = informant_names[0].strip()
        p.informant_surname = informant_names[-1].strip()
        p.informant_midname = ' '.join(informant_names[1:-1]).strip()

        if len(p.dri)<9:
            p.dri = 'FAILED TO READ'
            error_list.append(f'{index+1}')

        if len(p.qualification.split(' '))>1:
            p.qualification = "informant"

        print(f'--------------------------------------------------')
        for items in vars(p).items():
            x = items[0]
            y = items[1]
            print(f'{x} : {y}')

        end_time = time.time()
        print(f'Time elapsed: {round(end_time-start_time)}s')

    except AttributeError:
        error_list.append(f'{index+1}')
        for items in vars(p).keys():
            setattr(p,items,"")

    for key,value in enumerate(fields):
        ws.cell(row=1,column=key+1,value=value.upper())
    for place,column in pos_data.items():
        ws.cell(row=index+2,column=column,value=getattr(p,place).upper())
    time.sleep(2)

print(f'--------------------------------------------------')
os.system('cls')
title()

try:
    wb.save('data.xlsx')

    for file_ in os.listdir():
        if file_.endswith('_pdfocr.csv'):
            os.remove(file_)
            print(f'deleted {file_}')

    print('converting to csv')
    workbook = load_workbook('data.xlsx',read_only=True,data_only=True)
    worksheets = workbook.sheetnames
    sheets = [sheet for sheet in worksheets]

    for index,sheet in enumerate(sheets):
        output_csv = open(f'{sheet}_pdfocr.csv','w',newline='')
        worksheet = workbook[sheet]

    wr = csv.writer(output_csv,quoting=csv.QUOTE_ALL)

    for row in worksheet.iter_rows():
        csv_row = []
        for cell in row:
            csv_row.append(cell.value)
        wr.writerow(csv_row)

    print('csv converted, cleaning up')
    output_csv.close()
    workbook.close()
    os.remove('data.xlsx')

except PermissionError:
    print('failed to save excel sheet, ensure it is not open')

print(f'--------------------------------------------------')
if len(error_list)>0:    
    print('pages for manual assessment:')
    print(f"{','.join(error_list)}")
print(f'--------------------------------------------------')
print('exiting in 5s')
time.sleep(5)
sys.exit()